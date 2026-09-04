from __future__ import annotations

import json
import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from app import main
from app.models import PermitRecord
from app.services.exporter import build_export


class ExportExcelOnlyTests(unittest.TestCase):
    def test_excel_has_extension_date_column_with_business_label(self) -> None:
        record = PermitRecord(
            id=1,
            permit_number="34567",
            device_id="device-1",
            worker_name="Иванов И.И.",
            structural_unit="ЦДПН-1",
            data_json=json.dumps(
                {
                    "BE": {
                        "stage_label": "Продление РПО",
                        "field_value": "16.08.2026",
                        "comment": "",
                    }
                },
                ensure_ascii=False,
            ),
            first_received_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc),
        )

        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path, _json_path = build_export([record], tmp, batch_id=91)
            ws = load_workbook(xlsx_path, read_only=True).active
            headers = [cell.value for cell in ws[1]]
            self.assertIn("BE — Продление работ на", headers)
            extension_col = headers.index("BE — Продление работ на") + 1
            self.assertEqual(ws.cell(row=2, column=extension_col).value, "16.08.2026")

    def test_email_export_forwards_only_xlsx_attachment(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = Path(tmp) / "RPO_UPDATE_1.xlsx"
            json_path = Path(tmp) / "RPO_UPDATE_1.json"
            xlsx_path.write_bytes(b"xlsx")
            json_path.write_text("{}", encoding="utf-8")

            with patch("app.main._original_send_export", return_value="resend:test") as mocked:
                result = main._send_xlsx_export_only(
                    "operator@example.ru",
                    "РПО — выгрузка",
                    "Тест",
                    [xlsx_path, json_path],
                    idempotency_key="rpo-export-1",
                )

            self.assertEqual(result, "resend:test")
            mocked.assert_called_once_with(
                "operator@example.ru",
                "РПО — выгрузка",
                "Тест",
                [xlsx_path],
                idempotency_key="rpo-export-1",
            )


if __name__ == "__main__":
    unittest.main()

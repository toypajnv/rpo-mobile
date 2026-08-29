from __future__ import annotations

import base64
import inspect
import json
import tempfile
import unittest
from datetime import datetime, timezone, timedelta
from email import policy
from email.parser import BytesParser
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy import func, select

from app.config import settings
from app.database import Base, SessionLocal, engine
from app.main import ensure_permit_records, mobile_history, _record_state, _analytics_context, _work_view, export_preview, send_export_confirmed
from app.models import MobileEvent, PermitRecord
from app.services.exporter import build_export
from app.services.mailer import send_export


class PermitRegressionTests(unittest.TestCase):
    def setUp(self) -> None:
        Base.metadata.drop_all(engine)
        Base.metadata.create_all(engine)

    def tearDown(self) -> None:
        Base.metadata.drop_all(engine)

    def _event(self, key: str, value: str, client_id: str) -> MobileEvent:
        labels = {
            "AT": "Начало подготовки",
            "AU": "Окончание подготовки",
            "AV": "Передача ОП к ОБПР",
            "AY": "Фактическое начало работ",
            "AZ": "Остановка работ",
            "BA": "Возобновление работ",
            "BC": "Окончание работ",
            "BE": "Продление РПО",
        }
        return MobileEvent(
            client_event_id=client_id,
            device_id="device-test-1",
            worker_name="Иванов И.И.",
            permit_number="34567",
            field_key=key,
            stage_label=labels[key],
            event_time=datetime.now(timezone.utc),
            field_value=value,
            comment="",
        )

    def _record(self, record_id: int, permit: str, worker: str, fields: dict) -> PermitRecord:
        now = datetime.now(timezone.utc)
        return PermitRecord(
            id=record_id,
            permit_number=permit,
            device_id=f"device-{record_id}",
            worker_name=worker,
            data_json=json.dumps(fields, ensure_ascii=False),
            first_received_at=now,
            updated_at=now,
        )

    def test_backfill_groups_many_events_into_one_permit_row(self) -> None:
        events = [
            self._event("AT", "15.08.2026 11:31", "test-event-at"),
            self._event("AU", "15.08.2026 11:31", "test-event-au"),
            self._event("AV", "15.08.2026 11:32", "test-event-av"),
            self._event("AY", "15.08.2026 11:32", "test-event-ay"),
            self._event("BC", "15.08.2026 11:32", "test-event-bc"),
        ]
        with SessionLocal() as db:
            db.add_all(events)
            db.commit()

        ensure_permit_records()

        with SessionLocal() as db:
            count = db.scalar(select(func.count()).select_from(PermitRecord))
            self.assertEqual(count, 1)
            record = db.scalar(select(PermitRecord).where(PermitRecord.permit_number == "34567"))
            self.assertIsNotNone(record)
            data = json.loads(record.data_json)
            self.assertEqual(set(data), {"AT", "AU", "AV", "AY", "BC"})
            self.assertEqual(data["BC"]["field_value"], "15.08.2026 11:32")

    def test_mobile_history_returns_one_item_per_permit(self) -> None:
        with SessionLocal() as db:
            db.add_all([
                self._event("AT", "15.08.2026 11:31", "history-event-at"),
                self._event("AU", "15.08.2026 11:35", "history-event-au"),
            ])
            db.commit()
            result = mobile_history(device_id="device-test-1", limit=30, db=db)

        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["permit_number"], "34567")
        self.assertEqual(result[0]["field_key"], "НД")
        self.assertIn("Начало подготовки", result[0]["field_value"])
        self.assertIn("Окончание подготовки", result[0]["field_value"])

    def test_work_state_and_analytics_are_derived_from_permit_fields(self) -> None:
        stopped = self._record(1, "10001", "Иванов", {
            "AY": {"field_value": "15.08.2026 10:00", "event_time": "2026-08-15T05:00:00+00:00"},
            "AZ": {"field_value": "15.08.2026 11:00", "event_time": "2026-08-15T06:00:00+00:00"},
        })
        completed = self._record(2, "10002", "Петров", {
            "AY": {"field_value": "15.08.2026 08:00", "event_time": "2026-08-15T03:00:00+00:00"},
            "BC": {"field_value": "15.08.2026 12:00", "event_time": "2026-08-15T07:00:00+00:00"},
        })
        extended = self._record(3, "10003", "Иванов", {
            "AY": {"field_value": "15.08.2026 09:00", "event_time": "2026-08-15T04:00:00+00:00"},
            "BE": {"field_value": "16.08.2026", "event_time": "2026-08-15T06:30:00+00:00"},
        })

        self.assertEqual(_record_state(stopped), ("Остановлено", "stopped"))
        self.assertEqual(_record_state(completed), ("Завершено", "done"))
        self.assertEqual(_record_state(extended), ("Продлено", "extended"))

        analytics = _analytics_context([stopped, completed, extended], [], datetime(2026, 8, 15, 12, 0, tzinfo=timezone.utc))
        self.assertEqual(analytics["total"], 3)
        self.assertEqual(analytics["stopped"], 1)
        self.assertEqual(analytics["completed"], 1)
        self.assertEqual(analytics["active"], 1)
        self.assertEqual(analytics["extended"], 1)
        self.assertEqual(analytics["avg_completion_hours"], 4.0)
        self.assertEqual(analytics["avg_completion_label"], "4.0 ч")
        self.assertEqual(analytics["top_workers"][0], {"name": "Иванов", "count": 2})
        self.assertTrue(all(item["key"] != "AX" for item in analytics["stage_progress"]))

    def test_work_progress_uses_only_operator_visible_stages(self) -> None:
        record = self._record(1, "34567", "Иванов", {
            "AT": {"field_value": "1"},
            "AU": {"field_value": "2"},
            "AV": {"field_value": "3"},
            "AY": {"field_value": "4"},
            "BC": {"field_value": "5"},
            "AX": {"field_value": "legacy hidden stage"},
        })
        view = _work_view(record)
        self.assertEqual(view["stage_count"], 5)
        self.assertEqual(view["stage_total"], 8)
        self.assertEqual({item["key"] for item in view["stage_items"]}, {"AT", "AU", "AV", "AY", "BC"})
        self.assertNotIn("AX", {item["key"] for item in view["stage_items"]})

    def test_export_preview_includes_previously_exported_permit(self) -> None:
        now = datetime.now(timezone.utc)
        record = self._record(1, "34567", "Иванов", {
            "AT": {"field_value": "15.08.2026 11:31", "comment": ""},
        })
        record.exported_at = now - timedelta(minutes=5)
        with SessionLocal() as db:
            db.add(record)
            db.commit()
            preview = export_preview(
                period_from=(now - timedelta(hours=1)).isoformat(),
                period_to=(now + timedelta(hours=1)).isoformat(),
                operator=None,
                db=db,
            )

        self.assertEqual(len(preview["records"]), 1)
        self.assertEqual(preview["records"][0]["permit_number"], "34567")
        self.assertTrue(preview["records"][0]["previously_exported"])
        self.assertEqual(preview["stage_keys"], ["AT", "AU", "AV", "AY", "AZ", "BA", "BE", "BC", "RI"])

        send_source = inspect.getsource(send_export_confirmed)
        self.assertNotIn("PermitRecord.exported_at.is_(None)", send_source)
        self.assertNotIn("export=duplicate", send_source)

    def _records(self) -> list[PermitRecord]:
        return [
            self._record(1, "34567", "Иванов И.И.", {"AT": {"field_value": "15.08.2026 11:31", "comment": "", "stage_label": "Начало подготовки"}}),
            self._record(2, "98765", "Петров П.П.", {"AY": {"field_value": "15.08.2026 12:00", "comment": "", "stage_label": "Фактическое начало работ"}}),
        ]

    def test_export_is_one_row_per_permit_and_one_file_email(self) -> None:
        old_mode = settings.mail_mode
        old_outbox = settings.outbox_dir
        try:
            with tempfile.TemporaryDirectory() as tmp:
                export_dir = Path(tmp) / "exports"
                outbox_dir = Path(tmp) / "outbox"
                xlsx_path, json_path = build_export(self._records(), str(export_dir), batch_id=77)

                ws = load_workbook(xlsx_path, read_only=True).active
                self.assertEqual(ws.max_row, 3)
                payload = json.loads(json_path.read_text(encoding="utf-8"))
                self.assertEqual(len(payload["permits"]), 2)

                settings.mail_mode = "file"
                settings.outbox_dir = str(outbox_dir)
                send_export(
                    "operator@example.ru",
                    "РПО — тест",
                    "Тест одной выгрузки одним письмом",
                    [xlsx_path, json_path],
                )

                eml_files = list(outbox_dir.glob("*.eml"))
                self.assertEqual(len(eml_files), 1)
                message = BytesParser(policy=policy.default).parsebytes(eml_files[0].read_bytes())
                self.assertEqual(message["To"], "operator@example.ru")
                self.assertEqual(len(list(message.iter_attachments())), 2)
        finally:
            settings.mail_mode = old_mode
            settings.outbox_dir = old_outbox

    def test_resend_api_uses_one_https_request_with_two_attachments(self) -> None:
        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

            def read(self):
                return b'{"id":"email-test-123"}'

        old_mode = settings.mail_mode
        old_key = settings.resend_api_key
        old_from = settings.resend_from
        old_url = settings.resend_api_url
        try:
            with tempfile.TemporaryDirectory() as tmp:
                xlsx_path, json_path = build_export(self._records(), str(Path(tmp) / "exports"), batch_id=88)
                settings.mail_mode = "resend"
                settings.resend_api_key = "re_test_secret"
                settings.resend_from = "РПО Сервер <rpo@rpo-mng.ru>"
                settings.resend_api_url = "https://api.resend.com/emails"

                with patch("app.services.mailer.urllib_request.urlopen", return_value=FakeResponse()) as mocked:
                    result = send_export(
                        "operator@example.ru",
                        "РПО — выгрузка №88",
                        "Одна выгрузка одним письмом",
                        [xlsx_path, json_path],
                        idempotency_key="rpo-export-88",
                    )

                self.assertEqual(result, "resend:email-test-123")
                self.assertEqual(mocked.call_count, 1)
                req = mocked.call_args.args[0]
                self.assertEqual(req.full_url, "https://api.resend.com/emails")
                headers = {k.lower(): v for k, v in req.header_items()}
                self.assertEqual(headers["authorization"], "Bearer re_test_secret")
                self.assertEqual(headers["idempotency-key"], "rpo-export-88")
                self.assertIn("rpo-server", headers["user-agent"].lower())

                body = json.loads(req.data.decode("utf-8"))
                self.assertEqual(body["from"], "РПО Сервер <rpo@rpo-mng.ru>")
                self.assertEqual(body["to"], ["operator@example.ru"])
                self.assertEqual(len(body["attachments"]), 2)
                filenames = {item["filename"] for item in body["attachments"]}
                self.assertEqual(filenames, {xlsx_path.name, json_path.name})
                for item in body["attachments"]:
                    decoded = base64.b64decode(item["content"])
                    original = xlsx_path if item["filename"] == xlsx_path.name else json_path
                    self.assertEqual(decoded, original.read_bytes())
        finally:
            settings.mail_mode = old_mode
            settings.resend_api_key = old_key
            settings.resend_from = old_from
            settings.resend_api_url = old_url


class DashboardStaticTests(unittest.TestCase):
    def test_operator_tabs_preview_and_cache_busting(self) -> None:
        server_dir = Path(__file__).resolve().parents[1]
        template = (server_dir / "app/templates/dashboard.html").read_text(encoding="utf-8")
        script = (server_dir / "app/static/dashboard.js").read_text(encoding="utf-8")

        for element_id in (
            'id="tab-transmissions"',
            'id="tab-works"',
            'id="tab-analytics"',
            'id="tab-settings"',
            'id="export-form"',
            'id="preview-modal"',
            'id="preview-list"',
            'id="confirm-export-form"',
            'id="confirm-send"',
        ):
            self.assertIn(element_id, template)
        self.assertIn('href="#settings"', template)
        self.assertIn('href="#analytics"', template)
        self.assertIn('data-tab-link="transmissions"', template)
        self.assertIn('data-tab-link="works"', template)
        self.assertIn("Один наряд-допуск — одна строка", template)
        self.assertIn("Ранее выгруженные НД тоже можно", template)
        self.assertIn("dashboard.js?v=20260829-2", template)
        self.assertIn("app.css?v=20260829-3", template)
        self.assertIn("preview-table", script)
        self.assertIn("stageDetails", script)
        self.assertIn("Показать детали", script)
        self.assertIn("За выбранный период и текущий фильтр нет нарядов-допусков.", script)
        self.assertNotIn("нет невыгруженных нарядов-допусков", script)
        self.assertIn("window.scrollTo", script)
        self.assertIn("exportForm?.addEventListener('submit'", script)
        self.assertIn("modal.hidden=false", script.replace(" ", ""))
        self.assertIn("/api/operator/transmissions", script)
        self.assertIn("/api/operator/events", script)


if __name__ == "__main__":
    unittest.main()
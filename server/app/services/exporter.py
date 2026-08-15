from __future__ import annotations
from datetime import datetime, timezone
from pathlib import Path
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ..models import PermitRecord
from ..stages import STAGES


def local_fmt(dt: datetime | None) -> str:
    if not dt:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone().strftime("%d.%m.%Y %H:%M:%S")


def record_data(record: PermitRecord) -> dict:
    try:
        value = json.loads(record.data_json or "{}")
        return value if isinstance(value, dict) else {}
    except Exception:
        return {}


def build_export(records: list[PermitRecord], export_dir: str, batch_id: int) -> tuple[Path, Path]:
    """Build one export row per permit (plus nested JSON for local import)."""
    out = Path(export_dir)
    out.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = out / f"RPO_UPDATE_{batch_id}_{stamp}.xlsx"
    json_path = out / f"RPO_UPDATE_{batch_id}_{stamp}.json"

    stage_keys = [key for key, _ in sorted(STAGES.items(), key=lambda item: item[1]["order"])]

    wb = Workbook()
    ws = wb.active
    ws.title = "Наряды-допуски"
    headers = ["НД", "Работник"] + [f"{key} — {STAGES[key]['label']}" for key in stage_keys] + ["Комментарии", "Обновлено", "ID записи"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B3A73")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    json_rows = []
    for record in records:
        data = record_data(record)
        comments = []
        values = []
        normalized_fields = {}
        for key in stage_keys:
            field = data.get(key) or {}
            value = str(field.get("field_value", ""))
            comment = str(field.get("comment", "")).strip()
            values.append(value)
            if comment:
                comments.append(f"{key}: {comment}")
            if field:
                normalized_fields[key] = {
                    "label": field.get("stage_label") or STAGES[key]["label"],
                    "value": value,
                    "event_time": field.get("event_time", ""),
                    "comment": comment,
                }

        ws.append([
            record.permit_number,
            record.worker_name,
            *values,
            "\n".join(comments),
            local_fmt(record.updated_at),
            record.id,
        ])
        json_rows.append({
            "record_id": record.id,
            "permit_number": record.permit_number,
            "worker_name": record.worker_name,
            "device_id": record.device_id,
            "updated_at": record.updated_at.isoformat(),
            "fields": normalized_fields,
        })

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    for idx in range(3, 3 + len(stage_keys)):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 25
    comments_col = 3 + len(stage_keys)
    ws.column_dimensions[ws.cell(row=1, column=comments_col).column_letter].width = 42
    ws.column_dimensions[ws.cell(row=1, column=comments_col + 1).column_letter].width = 22
    ws.column_dimensions[ws.cell(row=1, column=comments_col + 2).column_letter].width = 12
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(xlsx_path)

    json_path.write_text(
        json.dumps({"version": 2, "batch_id": batch_id, "permits": json_rows}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return xlsx_path, json_path

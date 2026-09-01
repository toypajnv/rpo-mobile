from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook
from sqlalchemy import DateTime, Integer, String, Text, delete, select
from sqlalchemy.orm import Mapped, Session, mapped_column

from .database import Base


VIDEO_MARKERS = ("СИДОРОВИЧ", "БДД")
DESCRIPTION_HEADER_MARKERS = (
    "НАРУШ",
    "ПРИЧИН",
    "ОПИС",
    "ОСНОВ",
    "КОММЕНТ",
    "ПРИМЕЧ",
    "СОБЫТИ",
    "ИНЦИДЕНТ",
)


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


class UkzBlockRecord(Base):
    __tablename__ = "ukz_block_records"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    fio_raw: Mapped[str] = mapped_column(String(300), default="")
    fio_normalized: Mapped[str] = mapped_column(String(300), default="", index=True)
    surname: Mapped[str] = mapped_column(String(120), default="", index=True)
    first_name: Mapped[str] = mapped_column(String(120), default="")
    patronymic: Mapped[str] = mapped_column(String(120), default="")
    block_kind: Mapped[str] = mapped_column(String(40), default="corporate", index=True)
    marker_text: Mapped[str] = mapped_column(Text, default="")
    violation_description: Mapped[str] = mapped_column(Text, default="")
    source_sheet: Mapped[str] = mapped_column(String(200), default="")
    source_row: Mapped[int] = mapped_column(Integer)
    imported_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=utcnow, index=True)


class UkzBlockImport(Base):
    __tablename__ = "ukz_block_imports"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    file_hash: Mapped[str] = mapped_column(String(64), unique=True, index=True)
    file_name: Mapped[str] = mapped_column(String(255), default="")
    message_id: Mapped[str] = mapped_column(String(500), default="", index=True)
    sender: Mapped[str] = mapped_column(String(320), default="")
    source_rows: Mapped[int] = mapped_column(Integer, default=0)
    indexed_rows: Mapped[int] = mapped_column(Integer, default=0)
    imported_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=utcnow, index=True)


@dataclass(frozen=True)
class NameParts:
    surname: str
    first_name: str
    patronymic: str


@dataclass(frozen=True)
class ParsedUkzRecord:
    fio_raw: str
    fio_normalized: str
    surname: str
    first_name: str
    patronymic: str
    block_kind: str
    marker_text: str
    violation_description: str
    source_sheet: str
    source_row: int


@dataclass(frozen=True)
class UkzImportResult:
    duplicate: bool
    source_rows: int
    indexed_rows: int
    file_hash: str


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm(value: object) -> str:
    text = _text(value).upper().replace("Ё", "Е")
    text = re.sub(r"[\u00a0\s]+", " ", text)
    return text.strip()


def _name_token(value: str) -> str:
    return re.sub(r"[^А-ЯA-Z-]", "", _norm(value))


def _name_tokens(value: object) -> list[str]:
    raw = _norm(value)
    if not raw:
        return []
    result: list[str] = []
    for raw_token in re.split(r"\s+", raw):
        token = _name_token(raw_token)
        if not token:
            continue
        # Corporate files often shorten first name and patronymic as "И.И.".
        # Preserve them as two independent initials rather than the token "ИИ".
        if "." in raw_token and 1 < len(token) <= 3 and "-" not in token:
            result.extend(list(token))
        else:
            result.append(token)
    return result


def parse_name(value: object) -> NameParts | None:
    tokens = _name_tokens(value)
    if not tokens:
        return None
    surname = tokens[0]
    first_name = tokens[1] if len(tokens) > 1 else ""
    patronymic = tokens[2] if len(tokens) > 2 else ""
    if len(surname) < 2:
        return None
    return NameParts(surname=surname, first_name=first_name, patronymic=patronymic)


def normalize_fio(value: object) -> str:
    parts = parse_name(value)
    if not parts:
        return ""
    return " ".join(part for part in (parts.surname, parts.first_name, parts.patronymic) if part)


def _header_text(value: object) -> str:
    return re.sub(r"[^А-ЯA-Z0-9]+", "", _norm(value))


def _is_fio_header(value: object) -> bool:
    compact = _header_text(value)
    return "ФИО" in compact or "ФИОРАБОТНИКА" in compact or "ФИОСОТРУДНИКА" in compact


def _sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _find_sheet_and_header(workbook) -> tuple[object, int, int, dict[int, str]]:
    best: tuple[int, object, int, int, dict[int, str]] | None = None
    for sheet in workbook.worksheets:
        max_scan = min(max(sheet.max_row, 1), 80)
        for row_index in range(1, max_scan + 1):
            row_values = [sheet.cell(row=row_index, column=col).value for col in range(1, sheet.max_column + 1)]
            fio_columns = [idx for idx, value in enumerate(row_values, start=1) if _is_fio_header(value)]
            if not fio_columns:
                continue
            fio_col = fio_columns[0]
            headers: dict[int, str] = {}
            for col in range(1, sheet.max_column + 1):
                fragments = []
                for header_row in range(max(1, row_index - 2), row_index + 1):
                    text = _norm(sheet.cell(row=header_row, column=col).value)
                    if text and text not in fragments:
                        fragments.append(text)
                headers[col] = " ".join(fragments)
            populated_after = 0
            for check_row in range(row_index + 1, min(sheet.max_row, row_index + 30) + 1):
                if _text(sheet.cell(row=check_row, column=fio_col).value):
                    populated_after += 1
            score = populated_after * 10 + sheet.max_column
            candidate = (score, sheet, row_index, fio_col, headers)
            if best is None or score > best[0]:
                best = candidate
    if best is None:
        raise ValueError('В стоп-листе УКЗ не найден столбец "ФИО"')
    _, sheet, header_row, fio_col, headers = best
    return sheet, header_row, fio_col, headers


def _description_columns(headers: dict[int, str], fio_col: int) -> list[int]:
    result: list[int] = []
    for col, header in headers.items():
        if col == fio_col:
            continue
        normalized = _norm(header)
        if any(marker in normalized for marker in DESCRIPTION_HEADER_MARKERS):
            result.append(col)
    return result


def _row_marker_text(values: list[object], fio_col: int) -> str:
    return " | ".join(
        _text(value)
        for col, value in enumerate(values, start=1)
        if col != fio_col and _text(value)
    )


def _is_video_marker(text: str) -> bool:
    normalized = _norm(text)
    return any(re.search(rf"(?<![А-ЯA-Z]){re.escape(marker)}(?![А-ЯA-Z])", normalized) for marker in VIDEO_MARKERS)


def _description_from_row(values: list[object], description_cols: list[int], fio_col: int) -> str:
    selected: list[str] = []
    for col in description_cols:
        value = _text(values[col - 1] if len(values) >= col else None)
        if value and value not in selected:
            selected.append(value)
    if selected:
        return " • ".join(selected)[:4000]

    fallback: list[str] = []
    for col, raw in enumerate(values, start=1):
        if col == fio_col:
            continue
        value = _text(raw)
        normalized = _norm(value)
        if not value or normalized in VIDEO_MARKERS or len(value) < 8:
            continue
        if value not in fallback:
            fallback.append(value)
    return " • ".join(fallback[:4])[:4000]


def parse_ukz_stoplist(path: str | Path) -> tuple[list[ParsedUkzRecord], int]:
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet, header_row, fio_col, headers = _find_sheet_and_header(workbook)
        description_cols = _description_columns(headers, fio_col)
        records: list[ParsedUkzRecord] = []
        source_rows = 0
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            values = list(row)
            fio_raw = _text(values[fio_col - 1] if len(values) >= fio_col else None)
            if not fio_raw:
                continue
            source_rows += 1
            parts = parse_name(fio_raw)
            if not parts or not parts.first_name:
                continue
            marker_text = _row_marker_text(values, fio_col)
            records.append(
                ParsedUkzRecord(
                    fio_raw=fio_raw,
                    fio_normalized=normalize_fio(fio_raw),
                    surname=parts.surname,
                    first_name=parts.first_name,
                    patronymic=parts.patronymic,
                    block_kind="video" if _is_video_marker(marker_text) else "corporate",
                    marker_text=marker_text[:4000],
                    violation_description=_description_from_row(values, description_cols, fio_col),
                    source_sheet=sheet.title,
                    source_row=row_index,
                )
            )
        if not records:
            raise ValueError("В стоп-листе УКЗ не найдено строк с ФИО для сопоставления")
        return records, source_rows
    finally:
        workbook.close()


def import_ukz_stoplist(
    db: Session,
    path: str | Path,
    *,
    file_name: str = "",
    message_id: str = "",
    sender: str = "",
) -> UkzImportResult:
    file_hash = _sha256_file(path)
    previous = db.scalar(select(UkzBlockImport).where(UkzBlockImport.file_hash == file_hash))
    if previous:
        return UkzImportResult(True, previous.source_rows, previous.indexed_rows, file_hash)

    parsed, source_rows = parse_ukz_stoplist(path)
    imported_at = utcnow()
    db.execute(delete(UkzBlockRecord))
    db.add_all(
        UkzBlockRecord(
            fio_raw=item.fio_raw,
            fio_normalized=item.fio_normalized,
            surname=item.surname,
            first_name=item.first_name,
            patronymic=item.patronymic,
            block_kind=item.block_kind,
            marker_text=item.marker_text,
            violation_description=item.violation_description,
            source_sheet=item.source_sheet,
            source_row=item.source_row,
            imported_at=imported_at,
        )
        for item in parsed
    )
    db.add(
        UkzBlockImport(
            file_hash=file_hash,
            file_name=file_name or Path(path).name,
            message_id=message_id[:500],
            sender=sender[:320],
            source_rows=source_rows,
            indexed_rows=len(parsed),
            imported_at=imported_at,
        )
    )
    db.commit()
    return UkzImportResult(False, source_rows, len(parsed), file_hash)


def _component_match(left: str, right: str) -> tuple[bool, int]:
    left = _name_token(left)
    right = _name_token(right)
    if not left or not right:
        return True, 0
    if left[0] != right[0]:
        return False, 0
    if len(left) == 1 or len(right) == 1:
        return True, 10
    if left == right:
        return True, 20
    return False, 0


def _match_score(target: NameParts, candidate: UkzBlockRecord) -> int | None:
    if target.surname != candidate.surname:
        return None
    first_ok, first_score = _component_match(target.first_name, candidate.first_name)
    if not first_ok or first_score == 0:
        return None
    patronymic_ok, patronymic_score = _component_match(target.patronymic, candidate.patronymic)
    if not patronymic_ok:
        return None
    return 100 + first_score + patronymic_score


def _join_descriptions(records: Iterable[UkzBlockRecord]) -> str:
    values: list[str] = []
    for record in records:
        value = (record.violation_description or "").strip()
        if value and value not in values:
            values.append(value)
    return " • ".join(values)[:4000]


def find_ukz_block(db: Session, fio: str) -> dict | None:
    target = parse_name(fio)
    if not target or not target.first_name:
        return None
    candidates = db.scalars(
        select(UkzBlockRecord).where(UkzBlockRecord.surname == target.surname)
    ).all()
    scored: list[tuple[int, UkzBlockRecord]] = []
    for candidate in candidates:
        score = _match_score(target, candidate)
        if score is not None:
            scored.append((score, candidate))
    if not scored:
        return None

    best_score = max(score for score, _ in scored)
    best = [record for score, record in scored if score == best_score]
    identities = {record.fio_normalized for record in best}

    if len(identities) > 1:
        return {
            "kind": "corporate",
            "description": "",
            "ambiguous": True,
        }

    kind = "video" if any(record.block_kind == "video" for record in best) else "corporate"
    return {
        "kind": kind,
        "description": _join_descriptions(best),
        "ambiguous": False,
    }
